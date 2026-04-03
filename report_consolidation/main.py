import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

df1 = pd.read_csv("reporte_enero.csv")
df2 = pd.read_csv("reporte_febrero.csv")
df3 = pd.read_csv("reporte_marzo.csv")

df_concat = pd.concat([df1, df2, df3]).reset_index(drop=True)

#Headers claros
df_concat.columns = ["Service Type", "Estudiante", "País", "Fecha" , "Fecha confirmada" , "Tipo" ,"Precio de la Clase $","Porcentaje del Profesor %", "Ganancia Neta $"]

#DataFrame definitivo y concatenado
df = df_concat[["Fecha", "Estudiante", "País", "Precio de la Clase $", "Tipo", "Porcentaje del Profesor %", "Ganancia Neta $"]]

###Tipos de clase
#Filtrar por Tipo de Clase
df = df[(df["Tipo"] == "Trial") | (df["Tipo"] == "Non-trial lesson")].reset_index(drop=True)

#Cambiar nombre de los Tipos de Clase
df["Tipo"] = df["Tipo"].replace({
    "Trial": "Prueba",
    "Non-trial lesson": "Regular",
})

###Paises
df["País"] = df["País"].replace({
    "Brazil": "Brasil",
    "United States": "Estados Unidos",
    "Philippines": "Filipinas",
    "Hungary": "Hungria",
})

###Porcentajes
#Transformar a entero y eliminar datos nulos
df["Porcentaje del Profesor %"] = pd.to_numeric(df["Porcentaje del Profesor %"], errors="coerce")
df["Porcentaje del Profesor %"] = df["Porcentaje del Profesor %"].fillna(0).astype(int)

###Ganancia Neta
df["Ganancia Neta $"] = pd.to_numeric(df["Ganancia Neta $"], errors="coerce")
df["Ganancia Neta $"] = df["Ganancia Neta $"].fillna(0).astype(float)

#Test/Debug
df.to_excel("reporte_trimestral_1.xlsx", index=False)

### Edición del documento con Openpyxl
###Metricas
# Asegurar que la columna 'fecha' sea datetime
df['Fecha'] = pd.to_datetime(df['Fecha'])

# 1. Ingresos totales (bruto)
ingresos_totales = df['Precio de la Clase $'].sum()

# 2. Ganancia neta total
ganancia_neta_total = df['Ganancia Neta $'].sum()

# 3. Clase promedio (neta)
clase_promedio = df['Ganancia Neta $'].mean()

# 4. Número total de clases
num_clases = len(df)

# 5. Mes con más clases (sin crear columna)
mes_con_mas_clases = df.groupby(df['Fecha'].dt.to_period('M')).size().idxmax()
clases_en_mes_max = df.groupby(df['Fecha'].dt.to_period('M')).size().max()

# 6. Alumno #1 (el que más ganancia neta generó)
alumno_top = df.groupby('Estudiante')['Ganancia Neta $'].sum().idxmax()
ganancia_alumno_top = df.groupby('Estudiante')['Ganancia Neta $'].sum().max()

# 7. País con más ingresos (ganancia neta)
pais_top = df.groupby('País')['Ganancia Neta $'].sum().idxmax()
ganancia_pais_top = df.groupby('País')['Ganancia Neta $'].sum().max()

# 8. Mes con mayor ingreso (sin crear columna)
mes_mayor_ingreso = df.groupby(df['Fecha'].dt.to_period('M'))['Ganancia Neta $'].sum().idxmax()
ingreso_mes_max = df.groupby(df['Fecha'].dt.to_period('M'))['Ganancia Neta $'].sum().max()

# 9. Número de clases de prueba y regulares
clases_prueba = df[df['Tipo'] == 'Prueba'].shape[0]

# 10. Regulares
clases_regulares = df[df['Tipo'] == 'Regular'].shape[0]

# Impresión de resultados
print("📊 RESUMEN DE KPI's")
print("="*40)
print(f"💰 Ingresos totales (bruto): ${ingresos_totales:,.2f}")
print(f"💵 Ganancia neta total: ${ganancia_neta_total:,.2f}")
print(f"📈 Clase promedio (neta): ${clase_promedio:,.2f}")
print(f"📚 Número total de clases: {num_clases}")
print(f"📅 Mes con más clases: {mes_con_mas_clases} ({clases_en_mes_max} clases)")
print(f"🏆 Alumno #1: {alumno_top} (${ganancia_alumno_top:,.2f} netos)")
print(f"🌍 País con más ingresos: {pais_top} (${ganancia_pais_top:,.2f} netos)")
print(f"📆 Mes con mayor ingreso: {mes_mayor_ingreso} (${ingreso_mes_max:,.2f} netos)")
print(f"🎓 Clases de prueba: {clases_prueba}")
print(f"📘 Clases regulares: {clases_regulares}")

#Abriendo el libro
wb = load_workbook("reporte_trimestral_1.xlsx")
ws = wb.active
ws.title = 'Detalles'

#Formateando la hoja de detalles
#Formato completo y alineacion
bold = Font(bold=True)
for celda in ws[1]:  # primera fila
    celda.font = bold
    celda.alignment = Alignment(horizontal='center')

for fila in ws.iter_rows(min_row=2):
    for celda in fila:  # primera fila
        celda.alignment = Alignment(horizontal='center')

# Ancho de columnas
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 15
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 25
ws.column_dimensions['G'].width = 18

print(df)

#Creación de la hoja de resumen
ws_resumen = wb.create_sheet('Resumen')

#Datos
resumen = [
    ['Concepto', 'Valor'],
    ['Ganancia neta total', f'${ganancia_neta_total:,.2f}'],
    ['Ingresos totales (bruto)', f'${ingresos_totales:,.2f}'],
    ['Número total de clases', f'{num_clases}'],
    ['Promedio neto por clase', f'${clase_promedio:,.2f}'],
    ['Clases regulares', f'{clases_regulares}'],
    ['Clases de prueba', f'{clases_prueba}'],
    ['Mes con más clases', f'{mes_con_mas_clases} ({clases_en_mes_max} clases)'],
    ['Mes con mayor ingreso', f'{mes_mayor_ingreso} (${ingreso_mes_max:,.2f} netos)'],
    ['Alumno #1', f'{alumno_top} (${ganancia_alumno_top:,.2f} netos)'],
    ['País con más ingresos', f'{pais_top} (${ganancia_pais_top:,.2f} netos)'],
]

for fila in resumen:
    ws_resumen.append(fila)

bold = Font(bold=True)
for celda in ws_resumen[1]:  # primera fila
    celda.font = bold
    celda.alignment = Alignment(horizontal='center')

for fila in ws_resumen.iter_rows(min_row=2):
    for celda in fila:  # primera fila
        celda.alignment = Alignment(horizontal='center')

ws_resumen.column_dimensions['A'].width = 25
ws_resumen.column_dimensions['B'].width = 30

#Guardar libro
wb.save("reporte_trimestral_1.xlsx")